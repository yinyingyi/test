from openpyxl import Workbook,load_workbook

class Practice():       #定义一个练习类
    def creat_data_one(self):       #定义一个creat()方法
        wb = Workbook()         #实例化Workbook()，alt+enter快捷导入
        ws1 = wb.active         # 获取当前页签
        ws1.title = "one"       # 给当前页签命名
        ws1["A1"] = "身高"        #在A1的位置输入"身高"
        ws1["B1"] = "体重"        #在B1的位置输入"体重"
        ws1["C1"] = "是否超重"      #在C1的位置输入"是否超重"
        data_dict = {180:81,160:51,175:71,158:41}   #定义字典
        data_keys = [i for i in data_dict.keys()]   #使用列表获取字典key值。要想单独获取字典每个值，必须先转化为列表
        for i in range(len(data_dict)):     #for循环，循环次数i=字典长度
            ws1.cell(row=i + 2, column=1).value = data_keys[i]      #写入字典key值。row行，col列，从excle第二行开始写值，所以row=i + 2
            ws1.cell(row=i + 2, column=2).value = data_dict[data_keys[i]]       #写入字典key对应的values值

        # 写法一：列表
        # height = [180,160,175,158]
        # weight = [80,50,70,40]
        # for i in range(len(height)):
        #     ws1.cell(row = i + 2,column=1).value = height[i]
        #     ws1.cell(row = i + 2,column=2).value = weight[i]
        wb.save("data_one.xlsx")

    def get_data(self):
        ld = load_workbook(filename="data_one.xlsx")
        sheet = ld["one"]
        for i in range(4):
            print(sheet.cell(row=i + 2, column=1).values)

    def health_data(self):
        ld = load_workbook(filename="data_one.xlsx")
        sheet = ld["one"]
        # height = [180, 160, 175, 158]
        # weight = [80,50,70,40]
        # heath_weight = (height - 70) *0.6
        for i in range(4):
            height = sheet.cell(row=i + 2, column=1).value
            weight = sheet.cell(row=i + 2, column=2).value
            mark = sheet.cell(row=i + 2, column=3).value
            heath_weight = (height - 70) *0.6
            if weight > heath_weight:
                print("超标")
                sheet.cell(row=i + 2, column=3).value="超标"
            elif weight == heath_weight:
                print("标准")
                sheet.cell(row=i + 2, column=3).value = "标准"
            else:
                print("太瘦")
                sheet.cell(row=i + 2, column=3).value = "太瘦"
            ld.save("data_one.xlsx")



p = Practice()
# p.creat_data_one()
# p.get_data()
p.health_data()
