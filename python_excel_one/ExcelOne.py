from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()     #实例化Workbook
dest_filename = 'empty_book.xlsx'   #定义文件名字
ws1 = wb.active     #获取当前页签
ws1.title = "range names"   #给当前页签命名
for row in range(1, 40):    #在第1-39行循环显示以下内容
    ws1.append(range(600))  #每行显示1-600
ws2 = wb.create_sheet(title="Pi")   #创建第二个sheet页
ws2['F5'] = 3.14    #在F5的位置添加3.14
ws3 = wb.create_sheet(title="Data") #创建第三个sheet页
for row in range(10, 20):       #在第10-19行循环显示以下内容
    for col in range(27, 54):       #在第27-53列循环显示以下内容
        # cell单元格，get_column_letter()列字母与数字之间的转换，{0}设置指定位置
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
print(ws3['AA10'].value)    #打印sheet3中AA10的值
wb.save(filename = dest_filename)   #保存文件
print()